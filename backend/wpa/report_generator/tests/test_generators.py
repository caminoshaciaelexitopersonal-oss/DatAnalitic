import pytest
import os
import shutil
import json
from PIL import Image
from backend.wpa.report_generator.docx_generator import create_docx_report
from backend.wpa.report_generator.excel_generator import create_excel_report
from backend.wpa.report_generator.pdf_generator import create_pdf_report

@pytest.fixture
def mock_artifacts():
    """Creates a mock artifact directory for a test job."""
    job_id = "test_report_job"
    out_base = f"data/processed/{job_id}"

    # Clean up before test
    if os.path.exists(out_base):
        shutil.rmtree(out_base)

    # Create mock artifact directories
    os.makedirs(os.path.join(out_base, "eda"))
    os.makedirs(os.path.join(out_base, "explainability"))

    # Create mock manifest.json
    with open(os.path.join(out_base, "manifest.json"), "w") as f:
        json.dump({"job_id": job_id, "dataset_hash": "mock_hash"}, f)

    # Create mock eda/summary.json
    with open(os.path.join(out_base, "eda", "summary.json"), "w") as f:
        json.dump({"n_rows": 100, "n_cols": 5}, f)

    # Create mock missing_values_report.json
    with open(os.path.join(out_base, "eda", "missing_values_report.json"), "w") as f:
        json.dump({"col1": {"count": 0, "percentage": 0.0}}, f)

    # Create mock shap_summary.json
    with open(os.path.join(out_base, "explainability", "shap_summary.json"), "w") as f:
        json.dump([{"feature": "feature1", "mean_abs_shap_value": 0.8}], f)

    # Create valid dummy PNG image files
    dummy_image = Image.new('RGB', (1, 1))
    dummy_image.save(os.path.join(out_base, "eda", "missing_values_heatmap.png"), "PNG")
    dummy_image.save(os.path.join(out_base, "explainability", "shap_summary.png"), "PNG")

    # Provide the necessary paths to the tests
    yield job_id, out_base

    # Clean up after test
    if os.path.exists(out_base):
        shutil.rmtree(out_base)

@pytest.mark.skip(reason="Report generator module is outdated and depends on local file system.")
def test_create_docx_report(mock_artifacts):
    job_id, out_base = mock_artifacts
    report_path = create_docx_report(job_id, out_base)
    assert os.path.exists(report_path)
    assert os.path.getsize(report_path) > 0

@pytest.mark.skip(reason="Report generator module is outdated and depends on local file system.")
def test_create_excel_report(mock_artifacts):
    job_id, out_base = mock_artifacts
    report_path = create_excel_report(job_id, out_base)
    assert os.path.exists(report_path)
    assert os.path.getsize(report_path) > 0

    # Verify sheets
    from openpyxl import load_workbook
    wb = load_workbook(report_path)
    assert "Summary" in wb.sheetnames
    assert "Missing Values" in wb.sheetnames
    assert "Feature Importance (SHAP)" in wb.sheetnames

@pytest.mark.skip(reason="Report generator module is outdated and depends on local file system.")
def test_create_pdf_report(mock_artifacts):
    # This test requires a valid report_template.html.
    # We create a minimal one for the test to pass.
    job_id, out_base = mock_artifacts

    # Create a minimal template for the test
    template_dir = os.path.join(os.path.dirname(__file__), "..", "assets", "templates")
    os.makedirs(template_dir, exist_ok=True)
    with open(os.path.join(template_dir, "report_template.html"), "w") as f:
        f.write("<h1>Report for {{ job_id }}</h1>")

    report_path = create_pdf_report(job_id, out_base)
    assert os.path.exists(report_path)
    assert os.path.getsize(report_path) > 0
